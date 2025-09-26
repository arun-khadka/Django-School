from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated

import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import get_object_or_404
from exam.utils.all_marksheet import get_all_students_marksheet_data
from exam.utils.marksheet_generator import generate_blank_marksheet
from rest_framework.parsers import MultiPartParser

from exam.utils.single_marksheet import get_single_student_marksheet_data
from school.models import School
from students.models import Student, StudentMarks, StudentSubjectMarks
from subject.models import Subject
from classes.models import Class
from section.models import Section
from django.http import HttpResponse
from .models import ExamTerm

from .serializers import ExamTermSerializer
from .serializers import MarksheetImportSerializer
from .utils.marksheet_importer import generate_marksheet_with_results


class ExamTermCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ExamTermSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Get the school from validated data
        school = serializer.validated_data.get("school")
        if not school:
            return Response(
                {"error": "School is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        # Ownership check
        if school.owner != request.user:
            return Response(
                {"error": "You do not own this school"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Save the term
        term = serializer.save()
        return Response(
            {
                "msg": "Exam term created successfully",
                "term": ExamTermSerializer(term).data,
            },
            status=status.HTTP_201_CREATED,
        )


class ExamTermListByUserView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Get all terms linked to schools owned by the logged-in user
        terms = ExamTerm.objects.filter(school__owner=request.user).select_related(
            "school"
        )

        serializer = ExamTermSerializer(terms, many=True)
        return Response(
            {
                "msg": "Exam terms retrieved successfully",
                "terms": serializer.data,
            },
            status=status.HTTP_200_OK,
        )


class ExamTermListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Fetch only terms for schools owned by logged-in user
        terms = ExamTerm.objects.filter(school__owner=request.user)
        serializer = ExamTermSerializer(terms, many=True)
        return Response(
            {"msg": "Exam terms retrieved successfully", "terms": serializer.data},
            status=status.HTTP_200_OK,
        )


class ExamTermListBySchoolView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, school_id):
        # Verify school ownership
        try:
            school = School.objects.get(id=school_id, owner=request.user)
        except School.DoesNotExist:
            return Response(
                {"error": "School not found or you do not own this school"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Fetch terms for this school
        terms = ExamTerm.objects.filter(school=school)
        serializer = ExamTermSerializer(terms, many=True)
        return Response(
            {"msg": "Exam terms retrieved successfully", "terms": serializer.data},
            status=status.HTTP_200_OK,
        )


class MarksheetExportView(APIView):
    """
    Export a blank marksheet for manual entry.
    Query params: school (id), class (id), section (id, optional), term (id)
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Get query parameters
            school_id = request.query_params.get("school")
            class_id = request.query_params.get("class")
            section_id = request.query_params.get("section")
            term_id = request.query_params.get("term")

            # Validate required params
            if not school_id or not class_id or not term_id:
                return Response(
                    {"error": "school, class, and term are required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Fetch school and check ownership
            school = get_object_or_404(School, id=school_id)
            if school.owner != request.user:
                return Response(
                    {"error": "You do not own this school"},
                    status=status.HTTP_403_FORBIDDEN,
                )

            # Fetch class and term
            class_obj = get_object_or_404(Class, id=class_id, school=school)
            term = get_object_or_404(ExamTerm, id=term_id, school=school)

            # Fetch section if provided
            section_obj = None
            if section_id:
                section_obj = get_object_or_404(Section, id=section_id, school=school)

            # Fetch subjects according to class and section
            subjects = Subject.objects.filter(class_obj=class_obj)
            if section_obj:
                subjects = subjects.filter(section=section_obj)
            else:
                subjects = subjects.filter(section__isnull=True)

            # Check if subjects exist
            if not subjects.exists():
                return Response(
                    {"error": "No subjects found for the selected class/section"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Generate Excel file
            return generate_blank_marksheet(
                school, class_obj, section_obj, term, subjects
            )

        except Exception as e:
            # Catch any unexpected error
            return Response(
                {"error": f"Internal server error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class ExamMarksheetGenerateView(APIView):
    """
    Generate an Excel marksheet for a specific term, class, and optional section.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, term_id, class_id, section_id=None):
        # Fetch exam term, ensuring it belongs to logged-in user's school
        term = get_object_or_404(ExamTerm, id=term_id, school__owner=request.user)

        # Fetch class and optional section
        class_obj = get_object_or_404(Class, id=class_id, school=term.school)

        section_obj = None
        if section_id:
            section_obj = get_object_or_404(Section, id=section_id, class_obj=class_obj)

        # Get subjects for class/section
        subjects = Subject.objects.filter(class_obj=class_obj)
        if section_obj:
            subjects = subjects.filter(section=section_obj)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{term.name} Marksheet"

        # Header row
        headers = ["S.N.", "Name", "Roll No", "OTP"] + [sub.name for sub in subjects]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        # Pre-fill 50 empty rows for manual entry
        for row_num in range(2, 52):
            ws[f"A{row_num}"] = row_num - 1  # SN auto-fill
            ws[f"D{row_num}"] = ""  # OTP column left blank for admin to fill

        # Send file as response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"{term.name}_{class_obj.grade}"
        if section_obj:
            filename += f"_{section_obj.name}"
        filename += "_marksheet.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class MarksheetImportView(APIView):
    """
    Import an Excel marksheet, save student data & marks to DB,
    and return a styled Excel with results & ranks.
    """

    def post(self, request):
        serializer = MarksheetImportSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        file = serializer.validated_data["file"]
        full_mark = serializer.validated_data.get("full_mark", 100)
        pass_mark = serializer.validated_data.get("pass_mark", 35)

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # Read metadata
            school_name = (
                str(ws.cell(row=1, column=1).value or "").replace("School:", "").strip()
            )
            class_row_value = str(ws.cell(row=2, column=1).value or "")
            parts = [p.strip() for p in class_row_value.split("|")]
            class_name = section_name = term_name = None
            for part in parts:
                if part.startswith("Class:"):
                    class_name = part.replace("Class:", "").strip()
                elif part.startswith("Section:"):
                    section_name = part.replace("Section:", "").strip()
                elif part.startswith("Term:"):
                    term_name = part.replace("Term:", "").strip()

            if not all([school_name, class_name, section_name, term_name]):
                return Response(
                    {"error": "School, Class, Section, or Term not found in Excel."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Fetch DB objects
            school = School.objects.get(name__iexact=school_name)
            grade_int = int(class_name)
            class_obj = Class.objects.get(grade=grade_int, school=school)
            section = Section.objects.get(
                name__iexact=section_name, class_obj=class_obj
            )
            term = ExamTerm.objects.get(name__iexact=term_name, school=school)

            # Read headers and subjects
            header_row = 4
            headers = [cell.value for cell in ws[header_row]]
            ignore_columns = ["OTP", "TOTAL", "PERCENTAGE", "GRADE", "RESULT", "RANK"]
            subject_cols = []
            subject_names = []

            for idx, header in enumerate(headers[3:], start=4):
                if str(header).strip().upper() not in ignore_columns:
                    subject_cols.append(idx)
                    subject_names.append(str(header).strip())

            subjects = {
                s.name.lower(): s
                for s in Subject.objects.filter(class_obj=class_obj, section=section)
            }

            # Save student marks to DB
            for row_idx in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(row=row_idx, column=2).value
                roll_no = ws.cell(row=row_idx, column=3).value
                otp = ws.cell(row=row_idx, column=4).value if "OTP" in headers else None
                if not name or not roll_no:
                    continue

                student, _ = Student.objects.get_or_create(
                    school=school,
                    class_obj=class_obj,
                    section=section,
                    roll_no=roll_no,
                    defaults={"name": name, "otp": otp},
                )

                student_marks, _ = StudentMarks.objects.get_or_create(
                    student=student, defaults={"term": term}
                )

                total = 0
                count = 0
                overall_pass = True
                for col_idx, subj_name in zip(subject_cols, subject_names):
                    mark = ws.cell(row=row_idx, column=col_idx).value
                    try:
                        mark = float(mark)
                    except:
                        mark = 0
                    subj = subjects.get(subj_name.lower())
                    if subj:
                        StudentSubjectMarks.objects.update_or_create(
                            student_marks=student_marks,
                            subject=subj,
                            defaults={"marks_obtained": mark},
                        )
                    total += mark
                    count += 1
                    if mark < pass_mark:
                        overall_pass = False

                percentage = (total / (full_mark * count)) * 100 if count else 0
                student_marks.total_marks = total
                student_marks.percentage = round(percentage, 2)
                student_marks.grade = (
                    student_marks.calculate_grade()
                    if hasattr(student_marks, "calculate_grade") and overall_pass
                    else "-"
                )
                student_marks.result = "Pass" if overall_pass else "Fail"
                student_marks.save()

            # Generate styled Excel
            return generate_marksheet_with_results(
                file, full_mark, pass_mark, school, class_obj, section, term
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


# For fetching all students marksheet
class StudentMarksRetrieveView(APIView):
    """
    Retrieve student marks with subjects, total, percentage, grade, result, and rank
    filtered by school, class (grade), section, and term.
    """

    def get(self, request):
        # Expect query params: school_id, grade, section_name, term_id
        school_id = request.query_params.get("school_id")
        grade = request.query_params.get("grade")
        section_name = request.query_params.get("section_name")
        term_id = request.query_params.get("term_id")

        if not all([school_id, grade, section_name, term_id]):
            return Response(
                {"error": "school_id, grade, section_name, term_id are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Fetch related objects
            school = School.objects.get(id=school_id)
            class_obj = Class.objects.get(school=school, grade=grade)
            section = Section.objects.get(
                class_obj=class_obj, name__iexact=section_name
            )
            term = ExamTerm.objects.get(id=term_id, school=school)

            # Fetch students marks
            student_marks_qs = StudentMarks.objects.filter(
                student__class_obj=class_obj, student__section=section, term=term
            ).order_by("-total_marks")

            # Calculate rank
            rank_dict = {
                sm.student_id: idx + 1 for idx, sm in enumerate(student_marks_qs)
            }

            # Build response data
            data = []
            for sm in student_marks_qs:
                # Get subject marks
                try:
                    subject_marks = {
                        s.subject.name: s.marks_obtained for s in sm.subject_marks.all()
                    }
                except AttributeError:
                    subject_marks = {
                        s.subject.name: s.marks_obtained
                        for s in sm.studentsubjectmarks_set.all()
                    }

                data.append(
                    {
                        "student": sm.student.name,
                        "roll_no": sm.student.roll_no,
                        "school": school.name,
                        "class_name": class_obj.grade,
                        "section": section.name,
                        "term": term.name,
                        "total_marks": sm.total_marks,
                        "percentage": sm.percentage,
                        "grade": sm.grade if sm.result == "Pass" else "-",
                        "result": sm.result,
                        "rank": rank_dict.get(sm.student_id),
                        "subjects": subject_marks,
                    }
                )

            return Response(
                {
                    "msg": f"{len(data)} students processed successfully",
                    "data": data,
                },
                status=status.HTTP_200_OK,
            )

        except School.DoesNotExist:
            return Response(
                {"error": "School not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Class.DoesNotExist:
            return Response(
                {"error": "Class not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Section.DoesNotExist:
            return Response(
                {"error": "Section not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except ExamTerm.DoesNotExist:
            return Response(
                {"error": "Term not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


# For single student marksheet
class SingleStudentMarksRetrieveView(APIView):
    """
    Retrieve a single student's marksheet by school name, class (grade), roll_no, section, and OTP.
    """

    def post(self, request):
        school_name = request.data.get("school_name")
        grade = request.data.get("grade")
        roll_no = request.data.get("roll_no")
        section_name = request.data.get("section_name")
        otp = request.data.get("otp")

        if not all([school_name, grade, roll_no, section_name, otp]):
            return Response(
                {
                    "error": "school_name, grade, roll_no, section_name, and otp are required"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Find school by name (created by anyone)
            school = School.objects.get(name__iexact=school_name.strip())

            # Find student matching all details
            student = Student.objects.get(
                school=school,
                class_obj__grade=grade,
                section__name__iexact=section_name.strip(),
                roll_no=roll_no,
                otp=otp,
            )

            # Get student marks
            student_marks = StudentMarks.objects.filter(student=student).first()
            if not student_marks:
                return Response(
                    {"error": "Marks not found for this student"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Calculate rank among peers in same school, class, section, and term
            peer_marks = StudentMarks.objects.filter(
                student__school=school,
                student__class_obj=student.class_obj,
                student__section=student.section,
                term=student_marks.term,
            ).order_by("-total_marks")
            rank_dict = {sm.student_id: idx + 1 for idx, sm in enumerate(peer_marks)}
            student_rank = rank_dict.get(student.id)

            # Subject-wise marks
            try:
                subject_marks = {
                    sm.subject.name: sm.marks_obtained
                    for sm in student_marks.subject_marks.all()
                }
            except AttributeError:
                subject_marks = {
                    sm.subject.name: sm.marks_obtained
                    for sm in StudentSubjectMarks.objects.filter(
                        student_marks=student_marks
                    )
                }

            data = {
                "student": student.name,
                "roll_no": student.roll_no,
                "school": student.school.name if student.school else None,
                "class_name": student.class_obj.grade if student.class_obj else None,
                "section": student.section.name if student.section else None,
                "term": student_marks.term.name if student_marks.term else None,
                "total_marks": student_marks.total_marks,
                "percentage": student_marks.percentage,
                "grade": student_marks.grade if student_marks.result == "Pass" else "-",
                "result": student_marks.result,
                "rank": student_rank,
                "subjects": subject_marks,
            }

            return Response(data, status=status.HTTP_200_OK)

        except School.DoesNotExist:
            return Response(
                {"error": "School not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Student.DoesNotExist:
            return Response(
                {"error": "No student found with provided details"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


# Fetches data only when files are uploaded
class AllMarksheetAPIView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")

        if not file:
            return Response(
                {"error": "File is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            data = get_all_students_marksheet_data(file)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(data, status=status.HTTP_200_OK)
