import openpyxl


def get_single_student_marksheet_data(file, otp_input, full_mark=100, pass_mark=35):
    """
    Returns a single student's marksheet as a dict based on OTP.
    Includes school name, class, section, term, and student marksheet and rank.
    FAIL if any subject is below pass_mark.
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # Extract school, class, section, term
    school_name = (
        str(ws.cell(row=1, column=1).value or "").replace("School:", "").strip()
    )
    class_name = str(ws.cell(row=2, column=1).value or "").replace("Class:", "").strip()
    section_name = (
        str(ws.cell(row=2, column=2).value or "").replace("Section:", "").strip()
    )
    term_name = str(ws.cell(row=2, column=3).value or "").replace("Term:", "").strip()

    header_row = 4
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row]]

    # Headers to exclude from subjects and from output
    exclude_headers = ["S.N.", "SN"]  # <-- skip SN completely
    ignore_for_subjects = [
        "S.N.",
        "SN",
        "Name",
        "Student Name",
        "Roll No.",
        "OTP",
        "Total",
        "Percentage",
        "Grade",
        "Result",
        "Rank",
    ]

    # Find OTP column
    try:
        otp_col = next(
            idx + 1 for idx, h in enumerate(headers) if h.strip().lower() == "otp"
        )
    except StopIteration:
        raise ValueError("OTP column not found in Excel")

    # Find student row by OTP
    student_row_idx = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if (
            str(ws.cell(row=row_idx, column=otp_col).value).strip()
            == str(otp_input).strip()
        ):
            student_row_idx = row_idx
            break

    if not student_row_idx:
        raise ValueError("No student found with this OTP")

    # Identify subject columns (exclude SN, Name, Roll, OTP, Total, Percentage, Grade, Result, Rank)
    subject_cols = [
        idx
        for idx, header in enumerate(headers, start=1)
        if header not in ignore_for_subjects
    ]

    total_marks = 0
    overall_pass = True
    student_data = {}

    for col_idx, header in enumerate(headers, start=1):
        if header in exclude_headers:
            continue  # skip SN completely

        value = ws.cell(row=student_row_idx, column=col_idx).value

        # Only process subject marks for calculation
        if col_idx in subject_cols:
            try:
                mark = int(value) if value not in [None, ""] else 0
            except:
                mark = 0

            total_marks += mark
            if mark < pass_mark:
                overall_pass = False

            student_data[header] = mark
        else:
            # Non-subject columns (Name, Roll, OTP, etc.)
            student_data[header] = value

    num_subjects = len(subject_cols)
    percentage = (
        round((total_marks / (full_mark * num_subjects)) * 100, 2)
        if num_subjects
        else 0
    )

    # Grade function
    def get_grade(pct):
        if pct >= 90:
            return "A+"
        elif pct >= 80:
            return "A"
        elif pct >= 70:
            return "B+"
        elif pct >= 60:
            return "B"
        elif pct >= 50:
            return "C+"
        elif pct >= 40:
            return "C"
        elif pct >= 33:
            return "D"
        else:
            return "F"

    student_data["Total"] = total_marks
    student_data["Percentage"] = percentage
    student_data["Grade"] = get_grade(percentage) if overall_pass else "-"
    student_data["Result"] = "PASS" if overall_pass else "FAIL"

    # Add school info
    result = {
        "school": school_name,
        "class": class_name,
        "section": section_name,
        "term": term_name,
        "student_marksheet": student_data,
    }

    return result
