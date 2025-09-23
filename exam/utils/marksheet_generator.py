import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl.utils import get_column_letter


def generate_blank_marksheet(school, class_obj, section, term, subjects):
    """
    Generate a styled blank marksheet Excel file for manual entry.
    Auto-fills SN and Roll No 1–20.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Safe sheet title
    sheet_title = f"{getattr(class_obj, 'grade', '')}-{getattr(term, 'name', '')}"
    for ch in ["[", "]", ":", "*", "?", "/", "\\"]:
        sheet_title = sheet_title.replace(ch, "")
    ws.title = sheet_title[:31] if sheet_title else "Marksheet"

    # Title rows (row 1–3)
    school_name = f"School: {getattr(school, 'name', '-')}"
    class_name = f"Class: {getattr(class_obj, 'grade', '-')}"
    section_name = f"Section: {getattr(section, 'name', '-') if section else '-'}"
    term_name = f"Term: {getattr(term, 'name', '-')}"

    # Fonts & alignment
    title_font = Font(bold=True, size=18, color="4F81BD")
    sub_title_font = Font(bold=True, size=12, color="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center")

    # Merge & style School Name row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(subjects) + 3)
    ws["A1"].value = school_name
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    # Merge & style Class/Section/Term row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(subjects) + 3)
    ws["A2"].value = f"{class_name} | {section_name} | {term_name}"
    ws["A2"].font = sub_title_font
    ws["A2"].alignment = center_align

    # Row 3 left empty for spacing
    ws["A3"].value = ""

    # Header row (row 4)
    headers = ["S.N.", "Student Name", "Roll No.", "OTP"] + [
        getattr(sub, "name", str(sub)) for sub in subjects
    ]
    ws.append(headers)

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Apply header styles
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Pre-fill 20 rows SN/Roll
    for i in range(1, 6):
        ws.append([i, "", i] + [""] * len(subjects))

    # Alternating row fill for readability
    fill_odd = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=5, max_row=9, min_col=1, max_col=len(headers)), start=5
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = fill_odd if row_idx % 2 else fill_even

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)  # FIXED
        for cell in col:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[column_letter].width = max_length + 2

    ws.column_dimensions["A"].width = 5  # SN small
    ws.column_dimensions["B"].width = 25  # Student Name bigger
    ws.column_dimensions["C"].width = 10  # Roll No smaller
    ws.column_dimensions["D"].width = 15  # OTP column bigger 👈


    # Freeze header row
    ws.freeze_panes = ws["A5"]
    # Save to response
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"marksheet_{getattr(class_obj, 'grade', '-')}_{getattr(term, 'name', '-')}.xlsx"
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
