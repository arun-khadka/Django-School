import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


def generate_marksheet_with_results(
    file, full_mark, pass_mark, school=None, class_obj=None, section=None, term=None
):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # === Read existing Excel info ===
    excel_school = (
        str(ws.cell(row=1, column=1).value or "").replace("School:", "").strip()
    )
    excel_class = (
        str(ws.cell(row=2, column=1).value or "").replace("Class:", "").strip()
    )

    # Prefer model attributes, fallback to Excel
    school_name = getattr(school, "name", None) or excel_school or "-"
    class_grade = (
        getattr(class_obj, "grade", getattr(class_obj, "name", None))
        or excel_class
        or "-"
    )
    section_name = getattr(section, "name", "-") if section else "-"
    term_name = getattr(term, "name", "-")

    # === Styles ===
    title_font = Font(bold=True, size=18, color="4F81BD")
    sub_title_font = Font(bold=True, size=12, color="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alt_fill_odd = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    alt_fill_even = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )

    # === Header row (assume row 4) ===
    header_row = 4
    headers = [cell.value for cell in ws[header_row]]

    # Identify subject columns, skip OTP if present
    subject_cols = []
    for idx, header in enumerate(headers[3:], start=4):  # columns 4 onward
        if str(header).strip().upper() not in [
            "OTP",
            "TOTAL",
            "PERCENTAGE",
            "GRADE",
            "RESULT",
            "RANK",
        ]:
            subject_cols.append(idx)

    num_subjects = len(subject_cols)
    last_subject_col = subject_cols[-1]  # last actual subject column

    # === Clean top rows ===
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row <= 3:
            ws.unmerge_cells(str(merge))

    for r in [1, 2, 3]:
        for c in range(1, last_subject_col + 6):  # extra columns for total/grade/rank
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # Row 1: School name
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=last_subject_col + 5
    )
    ws["A1"].value = f"School: {school_name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    # Row 2: Class + Section + Term
    ws.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=last_subject_col + 5
    )
    ws["A2"].value = (
        f"Class: {class_grade} | Section: {section_name} | Term: {term_name}"
    )
    ws["A2"].font = sub_title_font
    ws["A2"].alignment = center_align

    ws["A3"].value = ""  # spacing row

    # === Add new headers: Total, Percentage, Grade, Result, Rank ===
    ws.cell(row=header_row, column=last_subject_col + 1, value="Total")
    ws.cell(row=header_row, column=last_subject_col + 2, value="Percentage")
    ws.cell(row=header_row, column=last_subject_col + 3, value="Grade")
    ws.cell(row=header_row, column=last_subject_col + 4, value="Result")
    ws.cell(row=header_row, column=last_subject_col + 5, value="Rank")
    rank_col = last_subject_col + 5

    # Style headers
    for col_num in range(1, rank_col + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # === Grade function ===
    def get_grade(pct):
        if pct >= 90:
            return "A+"
        elif pct >= 80:
            return "A"
        elif pct >= 70:
            return "B+"
        elif pct >= 60:
            return "B"
        elif pct >= 50:
            return "C+"
        elif pct >= 40:
            return "C"
        elif pct >= 33:
            return "D"
        else:
            return "F"

    # === First pass: calculate totals ===
    student_totals = {}  # row_idx -> total
    for row_idx in range(header_row + 1, ws.max_row + 1):
        sn_cell = ws.cell(row=row_idx, column=1).value
        if not sn_cell or not str(sn_cell).isdigit():
            continue

        total_marks = 0
        for col_idx in subject_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                mark = int(cell.value) if cell.value not in [None, ""] else 0
            except:
                mark = 0
            total_marks += mark
        student_totals[row_idx] = total_marks

    # === Rank calculation ===
    sorted_totals = sorted(student_totals.items(), key=lambda x: x[1], reverse=True)
    rank_map = {}
    current_rank = 0
    last_total = None

    for idx, (row_idx, total) in enumerate(sorted_totals, start=1):
        if total != last_total:
            current_rank = idx
        rank_map[row_idx] = current_rank
        last_total = total

    # === Helper: Convert number to ordinal ===
    def ordinal(n):
        if 10 <= n % 100 <= 20:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
        return f"{n}{suffix}"

    # === Second pass: write results ===
    for row_idx, total_marks in student_totals.items():
        overall_pass = True

        # Subject marks with pass/fail coloring
        for col_idx in subject_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                mark = int(cell.value) if cell.value not in [None, ""] else 0
            except:
                mark = 0

            if mark < pass_mark:
                overall_pass = False
                cell.fill = fail_fill
            else:
                cell.fill = pass_fill

            cell.alignment = center_align
            cell.border = thin_border

        # Total
        total_cell = ws.cell(row=row_idx, column=last_subject_col + 1)
        total_cell.value = total_marks
        total_cell.alignment = center_align
        total_cell.border = thin_border

        # Percentage
        pct = round((total_marks / (full_mark * num_subjects)) * 100, 2)
        pct_cell = ws.cell(row=row_idx, column=last_subject_col + 2)
        pct_cell.value = pct
        pct_cell.alignment = Alignment(horizontal="left", vertical="center")
        pct_cell.border = thin_border

        # Grade
        grade_cell = ws.cell(row=row_idx, column=last_subject_col + 3)
        grade_cell.value = get_grade(pct) if overall_pass else "-"
        grade_cell.alignment = center_align
        grade_cell.border = thin_border

        # Result
        result_cell = ws.cell(row=row_idx, column=last_subject_col + 4)
        result_cell.value = "PASS" if overall_pass else "FAIL"
        result_cell.fill = pass_fill if overall_pass else fail_fill
        result_cell.alignment = center_align
        result_cell.font = Font(bold=True)
        result_cell.border = thin_border

        # Rank
        rank_cell = ws.cell(row=row_idx, column=last_subject_col + 5)
        rank_cell.value = ordinal(rank_map[row_idx])
        rank_cell.alignment = center_align
        rank_cell.border = thin_border

        # Alternating row fill for SN, Name, Roll
        row_fill = alt_fill_odd if row_idx % 2 else alt_fill_even
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = center_align

    # === Auto column widths ===
    for i, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col_cells:
            if isinstance(cell, MergedCell):
                continue
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    # Adjust SN and Name width
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25

    # Freeze headers
    ws.freeze_panes = ws[f"A{header_row + 1}"]

    # === Return Excel file ===
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="marksheet_result.xlsx"'
    return response
